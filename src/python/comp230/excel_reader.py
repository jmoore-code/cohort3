from openpyxl import load_workbook

wb = load_workbook(filename = "jm_invoice_data.xlsx")
sheet = wb.active

for value in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
    print(value)
